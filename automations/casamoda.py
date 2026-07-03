from __future__ import annotations

import csv
import json
import os
import re
import time
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from html import unescape
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import parse_qs, unquote, urljoin, urlparse

import openpyxl
import pandas as pd
import requests

from config import BASE_DIR, CASAMODA_PASSWORD, CASAMODA_USERNAME
from automations.supplier_profile import CASAMODA_VENTI_PROFILE


BASE_URL = "https://b2b.casamoda.com"
LOGIN_URL = f"{BASE_URL}/de/de/account/login"
AUTH_URL = f"{BASE_URL}/de/account/authenticate"
VENTI_MODERN_FIT_URL = (
    f"{BASE_URL}/de/de/article_collection/product-list-venti--modern-fit"
)
VENTI_AUTOIMPORT_URLS = (
    VENTI_MODERN_FIT_URL,
    f"{BASE_URL}/de/de/article_collection/product-list-venti--body-fit",
    f"{BASE_URL}/de/de/article_collection/product-list-venti--comfort-fit",
    f"{BASE_URL}/de/de/article_collection/product-list-venti--jerseyflex",
    f"{BASE_URL}/de/de/article_collection/product-list-venti-evening",
    f"{BASE_URL}/de/de/article_collection/product-list-venti--polos--shirts",
    f"{BASE_URL}/de/de/article_collection/product-list-venti--sakkos--westen",
)
CASAMODA_PRODUCT_FIELDS = (
    "sku",
    "Productnaam",
    "name",
    "supplier",
    "brand",
    "category",
    "article_number",
    "farbnummer",
    "color",
    "color_name",
    "color_missing",
    "fit",
    "collar",
    "design",
    "quality",
    "fabriccomp",
    "sleeve",
    "cuff",
    "sizes",
    "rrp",
    "variant_prices",
    "purchase_prices",
    "variant_stocks",
    "source_url",
    "source_description",
    "image_urls",
    "image_count",
    "has_images",
    "source_category_slug",
    "source_category_url",
    "magento_ready",
    "blocked_reason",
)


class UnknownPriceError(ValueError):
    pass


def _amount(value: object) -> Decimal:
    if value is None:
        raise UnknownPriceError("Missing price")

    text = str(value).strip()
    if not text:
        raise UnknownPriceError("Missing price")

    text = (
        text.replace("EUR", "")
        .replace("eur", "")
        .replace("\u20ac", "")
        .replace(" ", "")
        .replace("\xa0", "")
    )
    if re.fullmatch(r"\d+", text) and len(text) > 3:
        amount = Decimal(text) / Decimal("1000")
    else:
        text = text.replace(".", "").replace(",", ".") if "," in text else text
        try:
            amount = Decimal(text)
        except InvalidOperation as ex:
            raise UnknownPriceError(f"Invalid price: {value}") from ex
    return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _money(value: Decimal) -> str:
    return f"{value:.2f}"


def _clean_text(value: str) -> str:
    text = re.sub(r"<[^>]+>", " ", value)
    text = unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def _parse_attrs(tag: str) -> dict[str, str]:
    return {
        key.lower(): unescape(value)
        for key, value in re.findall(r"([\w:-]+)\s*=\s*['\"]([^'\"]*)['\"]", tag)
    }


def _json_dump(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


@dataclass
class PriceMiss:
    article_number: str
    farbnummer: str
    size: str
    purchase_price: str
    source_url: str


@dataclass
class CasamodaColorMiss:
    article_number: str
    farbnummer: str
    source_url: str


@dataclass(frozen=True)
class CasamodaColorRange:
    start: int
    end: int
    color: str


class CasamodaColorMap:
    DEFAULT_RANGES = (
        CasamodaColorRange(0, 99, "Wit"),
        CasamodaColorRange(100, 149, "Blauw"),
        CasamodaColorRange(150, 199, "Aqua tot petrol"),
        CasamodaColorRange(200, 299, "Bruin"),
        CasamodaColorRange(300, 349, "Groen"),
        CasamodaColorRange(350, 399, "Turquoise"),
        CasamodaColorRange(400, 449, "Rood"),
        CasamodaColorRange(450, 499, "Oranje"),
        CasamodaColorRange(500, 599, "Geel"),
        CasamodaColorRange(600, 699, "Beige"),
        CasamodaColorRange(700, 749, "Zilver"),
        CasamodaColorRange(750, 799, "Antraciet"),
        CasamodaColorRange(800, 899, "Zwart"),
        CasamodaColorRange(900, 949, "Lila"),
        CasamodaColorRange(950, 999, "Pruim"),
    )

    def __init__(self, ranges: Iterable[CasamodaColorRange]):
        self.ranges = tuple(ranges)

    @classmethod
    def default(cls) -> "CasamodaColorMap":
        return cls(cls.DEFAULT_RANGES)

    @classmethod
    def from_rows(cls, rows: Iterable[Iterable[object]]) -> "CasamodaColorMap":
        ranges: list[CasamodaColorRange] = []
        for row in rows:
            values = list(row)
            if len(values) < 3:
                continue
            try:
                start = int(str(values[0]).strip())
                end = int(str(values[1]).strip())
            except (TypeError, ValueError):
                continue

            color_value = values[3] if len(values) >= 4 and values[3] else values[2]
            color = str(color_value).strip()
            if not color:
                continue
            ranges.append(CasamodaColorRange(start, end, color))
        return cls(ranges)

    @classmethod
    def from_excel(cls, path: str | Path) -> "CasamodaColorMap":
        workbook = openpyxl.load_workbook(path, data_only=True)
        worksheet = workbook.active
        return cls.from_rows(worksheet.iter_rows(values_only=True))

    def lookup(self, farbnummer: str) -> str | None:
        try:
            code = int(str(farbnummer).strip())
        except (TypeError, ValueError):
            return None

        for color_range in self.ranges:
            if color_range.start <= code <= color_range.end:
                return color_range.color
        return None


class CasamodaPriceList:
    PROMPT_DEFAULTS = {
        Decimal("25.45"): "64.99",
        Decimal("27.75"): "69.99",
    }

    def __init__(self, mapping: dict[Decimal, str]):
        self.mapping = mapping

    @classmethod
    def from_rows(
        cls,
        rows: Iterable[tuple[object, object]],
        *,
        include_prompt_defaults: bool = True,
    ) -> "CasamodaPriceList":
        mapping: dict[Decimal, str] = {}
        for purchase, retail in rows:
            try:
                purchase_price = _amount(purchase)
                retail_price = _money(_amount(retail))
            except UnknownPriceError:
                continue
            mapping[purchase_price] = retail_price

        if include_prompt_defaults:
            for purchase_price, retail_price in cls.PROMPT_DEFAULTS.items():
                mapping.setdefault(purchase_price, retail_price)

        return cls(mapping)

    @classmethod
    def from_excel(cls, path: str | Path) -> "CasamodaPriceList":
        workbook = openpyxl.load_workbook(path, data_only=True)
        worksheet = workbook.active
        rows = (
            (row[0], row[1])
            for row in worksheet.iter_rows(values_only=True)
            if len(row) >= 2
        )
        return cls.from_rows(rows)

    def lookup(self, purchase_price: object) -> str:
        price = _amount(purchase_price)
        if price not in self.mapping:
            raise UnknownPriceError(f"Unknown Casamoda purchase price: {_money(price)}")
        return self.mapping[price]


class CasamodaParser:
    DETAIL_LABELS = (
        "Passform",
        "Armlänge",
        "Kragenform",
        "Muster",
        "Material",
        "Stoffart",
        "Farbe",
        "Größe",
    )

    DETAIL_STOP_LABELS = DETAIL_LABELS + ("NOS",)

    def __init__(
        self,
        price_list: CasamodaPriceList,
        *,
        color_map: CasamodaColorMap | None = None,
        missing_color_callback: Callable[[CasamodaColorMiss], None] | None = None,
    ):
        self.price_list = price_list
        self.color_map = color_map if color_map is not None else CasamodaColorMap.default()
        self.missing_color_callback = missing_color_callback
        self._reported_missing_colors: set[tuple[str, str, str]] = set()

    @staticmethod
    def parse_listing_links(html_text: str, base_url: str = BASE_URL) -> list[str]:
        links: set[str] = set()
        for href in re.findall(r"href\s*=\s*['\"]([^'\"]+/article/\d+/\d+[^'\"]*)", html_text):
            links.add(urljoin(base_url, href.split("#")[0]))
        return sorted(links)

    @staticmethod
    def parse_article_color_links(html_text: str, base_url: str = BASE_URL) -> list[str]:
        links: set[str] = set()
        for raw_url in re.findall(
            r"data-article-url\s*=\s*['\"]([^'\"]+)['\"]",
            html_text,
            re.IGNORECASE,
        ):
            clean_url = unescape(raw_url).split("#")[0]
            if re.search(r"/article/\d+/\d+", clean_url):
                links.add(urljoin(base_url, clean_url))
        return sorted(links)

    @staticmethod
    def group_article_urls(urls: Iterable[str]) -> list[str]:
        grouped: dict[str, str] = {}
        for url in urls:
            match = re.search(r"/article/(\d+)/\d+", url)
            article_key = match.group(1) if match else url
            grouped.setdefault(article_key, url)
        return list(grouped.values())

    def parse_product_detail(
        self,
        html_text: str,
        source_url: str,
        *,
        only_farbnummer: str | Iterable[str] | None = None,
    ) -> list[dict[str, str]]:
        article_number = self._article_number(html_text)
        details = self._product_details(html_text)
        description = self._description(html_text)
        page_name = self._page_name(html_text)
        rows: list[dict[str, str]] = []
        missing_prices: list[PriceMiss] = []
        missing_seen: set[tuple[str, str, str, str]] = set()
        selected_farbnummers: set[str] = set()
        if isinstance(only_farbnummer, str):
            selected_farbnummers = {only_farbnummer.strip()} if only_farbnummer.strip() else set()
        elif only_farbnummer:
            selected_farbnummers = {
                str(farbnummer).strip()
                for farbnummer in only_farbnummer
                if str(farbnummer).strip()
            }

        def add_missing(farbnummer: str, size: str, purchase_price: str) -> None:
            key = (article_number, farbnummer, size, purchase_price)
            if key in missing_seen:
                return
            missing_seen.add(key)
            missing_prices.append(
                PriceMiss(
                    article_number=article_number,
                    farbnummer=farbnummer,
                    size=size,
                    purchase_price=purchase_price,
                    source_url=source_url,
                )
            )

        variant_groups = self._variant_groups(html_text)
        if not variant_groups:
            variant_groups = [
                (farbnummer, color_name, self._variants(segment))
                for farbnummer, color_name, segment in self._color_segments(html_text)
            ]

        for farbnummer, color_name, variants in variant_groups:
            if selected_farbnummers and farbnummer not in selected_farbnummers:
                continue
            if not variants:
                continue

            sizes: list[str] = []
            purchase_prices: dict[str, str] = {}
            variant_prices: dict[str, str] = {}
            variant_stocks: dict[str, str] = {}

            for variant in variants:
                size = str(variant.get("data-item-size", "")).strip()
                if not size:
                    continue
                try:
                    purchase_price = _money(_amount(self._purchase_price_source(variant)))
                except UnknownPriceError:
                    add_missing(farbnummer, size, "missing")
                    continue
                try:
                    retail_price = self.price_list.lookup(purchase_price)
                except UnknownPriceError:
                    add_missing(farbnummer, size, purchase_price)
                    continue

                if size not in sizes:
                    sizes.append(size)
                purchase_prices[size] = purchase_price
                variant_prices[size] = retail_price
                variant_stocks[size] = str(variant.get("max", "")).strip()

            if not sizes:
                continue

            retail_values = [_amount(value) for value in variant_prices.values()]
            color_group, color_missing = self._color_group_for_row(
                article_number,
                farbnummer,
                source_url,
            )
            image_urls = self._image_urls_for_farbnummer(
                html_text,
                source_url,
                farbnummer,
            )
            quality = details.get("Stoffart", "")
            product_name = self._product_name(color_group, quality)
            sku = f"{article_number}-{farbnummer}"

            rows.append(
                {
                    "sku": sku,
                    "Productnaam": product_name,
                    "name": page_name,
                    "supplier": "casamoda",
                    "brand": "VENTI",
                    "category": "shirts",
                    "article_number": article_number,
                    "farbnummer": farbnummer,
                    "color": color_group,
                    "color_name": color_name,
                    "color_missing": color_missing,
                    "fit": details.get("Passform", ""),
                    "collar": details.get("Kragenform", ""),
                    "design": self._normalize_design(details.get("Muster", "")),
                    "quality": quality,
                    "fabriccomp": details.get("Material", ""),
                    "sleeve": details.get("Armlänge", ""),
                    "cuff": "SINGLE CUFF",
                    "sizes": str(sizes),
                    "rrp": _money(min(retail_values)),
                    "variant_prices": _json_dump(variant_prices),
                    "purchase_prices": _json_dump(purchase_prices),
                    "variant_stocks": _json_dump(variant_stocks),
                    "source_url": source_url,
                    "source_description": description,
                    "image_urls": _json_dump(image_urls),
                    "image_count": str(len(image_urls)),
                    "has_images": str(bool(image_urls)),
                }
            )

        if missing_prices:
            details_text = "; ".join(
                f"{miss.article_number}-{miss.farbnummer} size {miss.size}: {miss.purchase_price}"
                for miss in missing_prices
            )
            raise UnknownPriceError(details_text)

        return rows

    @classmethod
    def color_group_for_farbnummer(cls, farbnummer: str) -> str:
        return CasamodaColorMap.default().lookup(farbnummer) or "Meerkleurig"

    def _color_group_for_row(
        self,
        article_number: str,
        farbnummer: str,
        source_url: str,
    ) -> tuple[str, str]:
        color = self.color_map.lookup(farbnummer)
        if color:
            return color, "False"

        key = (article_number, farbnummer, source_url)
        if self.missing_color_callback is not None and key not in self._reported_missing_colors:
            self._reported_missing_colors.add(key)
            self.missing_color_callback(
                CasamodaColorMiss(
                    article_number=article_number,
                    farbnummer=farbnummer,
                    source_url=source_url,
                )
            )
        return "", "True"

    @classmethod
    def _article_number(cls, html_text: str) -> str:
        for pattern in (
            r"data-vendor-key\s*=\s*['\"]([^'\"]+)['\"]",
            r"Artikelnummer\s*:?\s*</?[^>]*>\s*([A-Za-z0-9-]+)",
            r"Artikelnummer\s*:?\s*([A-Za-z0-9-]+)",
        ):
            match = re.search(pattern, html_text, re.IGNORECASE)
            if match:
                return _clean_text(match.group(1))
        raise ValueError("Could not find Casamoda article number")

    @classmethod
    def _page_name(cls, html_text: str) -> str:
        match = re.search(r"<h1[^>]*>(.*?)</h1>", html_text, re.IGNORECASE | re.DOTALL)
        return _clean_text(match.group(1)) if match else "Businesshemd"

    @classmethod
    def _product_details(cls, html_text: str) -> dict[str, str]:
        details: dict[str, str] = {}
        label_pattern = "|".join(re.escape(label) for label in cls.DETAIL_LABELS)
        for match in re.finditer(
            r"<(?:li|p|span|td|th)[^>]*>(.*?)</(?:li|p|span|td|th)>",
            html_text,
            re.IGNORECASE | re.DOTALL,
        ):
            text = _clean_text(match.group(1))
            detail_match = re.match(
                rf"^({label_pattern})\s*:\s*(.+)$",
                text,
                re.IGNORECASE,
            )
            if detail_match:
                label = cls._detail_label(detail_match.group(1))
                details[label] = cls._clean_detail_value(detail_match.group(2))

        plain = _clean_text(html_text)
        lookahead = "|".join(re.escape(label) for label in cls.DETAIL_STOP_LABELS)
        for label in cls.DETAIL_LABELS:
            if label in details:
                continue
            pattern = rf"{re.escape(label)}\s*:\s*(.*?)(?=\s+(?:{lookahead})\s*:|Produktinformationen|Pflegehinweise|$)"
            match = re.search(pattern, plain, re.IGNORECASE)
            if match:
                details[label] = cls._clean_detail_value(match.group(1))
        return details

    @classmethod
    def _detail_label(cls, label: str) -> str:
        label_lower = label.strip().lower()
        for known_label in cls.DETAIL_LABELS:
            if known_label.lower() == label_lower:
                return known_label
        return label.strip()

    @staticmethod
    def _clean_detail_value(value: str) -> str:
        value = value.strip()
        value = re.split(r"\s+NOS\s*:", value, maxsplit=1, flags=re.IGNORECASE)[0]
        return value.strip()

    @classmethod
    def _description(cls, html_text: str) -> str:
        match = re.search(
            r"Produktinformationen(.*?)(?:Pflegehinweise|Herstellerinformation|$)",
            html_text,
            re.IGNORECASE | re.DOTALL,
        )
        return _clean_text(match.group(1)) if match else ""

    @classmethod
    def _image_urls(cls, html_text: str, source_url: str) -> list[str]:
        selected: dict[str, tuple[int, int, str]] = {}
        order = 0
        for attr in ("src", "data-src", "data-zoom-image"):
            pattern = rf"{attr}\s*=\s*['\"]([^'\"]+\.(?:jpg|jpeg|png|webp)(?:\?[^'\"]*)?)['\"]"
            for image_url in re.findall(pattern, html_text, re.IGNORECASE):
                absolute = urljoin(source_url, unescape(image_url))
                parsed_url = urlparse(absolute)
                query = parse_qs(parsed_url.query)
                if query.get("h") == ["50"] or query.get("w") == ["50"]:
                    continue

                key = parsed_url.path
                score = cls._image_url_score(parsed_url.query)
                current = selected.get(key)
                if current is None or score < current[0]:
                    selected[key] = (score, order, absolute)
                order += 1

        return [
            data[2]
            for data in sorted(selected.values(), key=lambda item: item[1])
        ]

    @classmethod
    def _image_urls_for_farbnummer(
        cls,
        html_text: str,
        source_url: str,
        farbnummer: str,
    ) -> list[str]:
        image_urls = cls._image_urls(html_text, source_url)
        matching_urls = [
            image_url
            for image_url in image_urls
            if cls._url_mentions_farbnummer(image_url, farbnummer)
        ]
        if matching_urls:
            return matching_urls

        selected_farbnummer = cls._selected_farbnummer(html_text, source_url)
        if selected_farbnummer == farbnummer:
            return image_urls

        swatch_urls = cls._swatch_image_urls_for_farbnummer(
            html_text,
            source_url,
            farbnummer,
        )
        if swatch_urls:
            return swatch_urls

        if selected_farbnummer and selected_farbnummer != farbnummer:
            return []

        return image_urls

    @classmethod
    def _swatch_image_urls_for_farbnummer(
        cls,
        html_text: str,
        source_url: str,
        farbnummer: str,
    ) -> list[str]:
        urls: list[str] = []
        for match in re.finditer(
            r"title\s*=\s*['\"]\s*(\d{3})\b[^'\"]*['\"]",
            html_text,
            re.IGNORECASE,
        ):
            if match.group(1) != farbnummer:
                continue
            end_candidates = [
                end_index
                for end_index in (
                    html_text.find("</a>", match.end()),
                    html_text.find("</div>", match.end()),
                )
                if end_index >= 0
            ]
            snippet_end = min(end_candidates) if end_candidates else match.start() + 800
            snippet = html_text[match.start() : snippet_end]
            raw_urls = re.findall(
                r"background-image\s*:\s*url\(([^)]+)\)",
                snippet,
                re.IGNORECASE,
            )
            raw_urls.extend(
                re.findall(
                    r"(?:src|data-src)\s*=\s*['\"]([^'\"]+\.(?:jpg|jpeg|png|webp)(?:\?[^'\"]*)?)['\"]",
                    snippet,
                    re.IGNORECASE,
                )
            )
            for raw_url in raw_urls:
                image_url = cls._normalize_swatch_image_url(raw_url, source_url)
                if image_url and image_url not in urls:
                    urls.append(image_url)
        return urls

    @staticmethod
    def _normalize_swatch_image_url(raw_url: str, source_url: str) -> str:
        clean_url = unescape(raw_url).strip().strip("'\"")
        if not clean_url:
            return ""
        absolute_url = urljoin(source_url, clean_url)
        parsed_url = urlparse(absolute_url)
        if not re.search(r"\.(?:jpg|jpeg|png|webp)$", parsed_url.path, re.IGNORECASE):
            return ""
        if parsed_url.query:
            return parsed_url._replace(query="auto=format").geturl()
        return absolute_url

    @staticmethod
    def _url_mentions_farbnummer(image_url: str, farbnummer: str) -> bool:
        code = str(farbnummer).strip()
        if not code:
            return False

        parsed_url = urlparse(image_url)
        url_text = unquote(parsed_url.path)
        return re.search(rf"(?<!\d){re.escape(code)}(?!\d)", url_text) is not None

    @staticmethod
    def _image_url_score(query: str) -> int:
        if query == "auto=format":
            return 0
        if "pad=5" in query:
            return 1
        if "pad=20" in query:
            return 2
        return 3

    @classmethod
    def _color_segments(cls, html_text: str) -> list[tuple[str, str, str]]:
        header_matches = list(
            re.finditer(
                r"<h[1-6][^>]*>\s*(\d{3}\s+.*?)</h[1-6]>",
                html_text,
                re.IGNORECASE | re.DOTALL,
            )
        )
        segments: list[tuple[str, str, str]] = []
        for index, match in enumerate(header_matches):
            label = _clean_text(match.group(1))
            label_match = re.match(r"(\d{3})\s+(.+?)(?:\s+SALE)?$", label, re.IGNORECASE)
            if not label_match:
                continue
            start = match.end()
            end = header_matches[index + 1].start() if index + 1 < len(header_matches) else len(html_text)
            segments.append((label_match.group(1), label_match.group(2).strip(), html_text[start:end]))

        if segments:
            return segments

        variants = cls._variants(html_text)
        if not variants:
            return []
        fallback_color = re.search(r"\b(\d{3})\s+([A-Za-zÀ-ÿ-]+)", _clean_text(html_text))
        if fallback_color:
            return [(fallback_color.group(1), fallback_color.group(2), html_text)]
        return [("000", "", html_text)]

    @staticmethod
    def _variants(segment: str) -> list[dict[str, str]]:
        variants: list[dict[str, str]] = []
        for tag in re.findall(r"<input\b[^>]*data-quantity[^>]*>", segment, re.IGNORECASE):
            attrs = _parse_attrs(tag)
            if not attrs.get("data-variant-id"):
                continue
            variants.append(attrs)
        return variants

    @classmethod
    def _variant_groups(cls, html_text: str) -> list[tuple[str, str, list[dict[str, str]]]]:
        color_titles = cls._color_titles(html_text)
        grouped: dict[str, list[dict[str, str]]] = {}
        for variant in cls._variants(html_text):
            farbnummer = cls._variant_farbnummer(variant)
            if not farbnummer:
                continue
            grouped.setdefault(farbnummer, []).append(variant)
        return [
            (farbnummer, color_titles.get(farbnummer, ""), variants)
            for farbnummer, variants in grouped.items()
        ]

    @staticmethod
    def _variant_farbnummer(variant: dict[str, str]) -> str:
        price_block = str(variant.get("data-article-price-block", ""))
        match = re.match(r"[^-]+-(\d{3})(?:-|$)", price_block)
        if match:
            return match.group(1)
        return ""

    @staticmethod
    def _color_titles(html_text: str) -> dict[str, str]:
        titles: dict[str, str] = {}
        for title in re.findall(r"title\s*=\s*['\"](\d{3}\s+[^'\"]+)['\"]", html_text, re.IGNORECASE):
            clean_title = _clean_text(title)
            match = re.match(r"(\d{3})\s+(.+?)(?:\s+SALE)?$", clean_title, re.IGNORECASE)
            if match:
                titles.setdefault(match.group(1), match.group(2).strip())
        return titles

    @classmethod
    def _selected_farbnummer(cls, html_text: str, source_url: str = "") -> str:
        for opening_tag, content in cls._color_bullet_blocks(html_text):
            classes = _parse_attrs(opening_tag).get("class", "").lower().split()
            if "checked" not in classes:
                continue
            farbnummer = cls._farbnummer_from_color_block(content)
            if farbnummer:
                return farbnummer

        if source_url:
            target_path = urlparse(source_url).path.rstrip("/")
            for _, content in cls._color_bullet_blocks(html_text):
                article_url = cls._color_block_article_url(content, source_url)
                if not article_url:
                    continue
                if urlparse(article_url).path.rstrip("/") != target_path:
                    continue
                farbnummer = cls._farbnummer_from_color_block(content)
                if farbnummer:
                    return farbnummer

        return ""

    @classmethod
    def _linked_farbnummer_codes(cls, html_text: str, source_url: str = BASE_URL) -> set[str]:
        farbnummers: set[str] = set()
        for _, content in cls._color_bullet_blocks(html_text):
            if not cls._color_block_article_url(content, source_url):
                continue
            farbnummer = cls._farbnummer_from_color_block(content)
            if farbnummer:
                farbnummers.add(farbnummer)
        return farbnummers

    @classmethod
    def _variant_farbnummer_codes(cls, html_text: str) -> set[str]:
        return {
            farbnummer
            for farbnummer, _, variants in cls._variant_groups(html_text)
            if variants
        }

    @staticmethod
    def _color_bullet_blocks(html_text: str) -> list[tuple[str, str]]:
        blocks: list[tuple[str, str]] = []
        for match in re.finditer(
            r"(<div\b[^>]*class\s*=\s*['\"][^'\"]*\bcolor-bullet\b[^'\"]*['\"][^>]*>)",
            html_text,
            re.IGNORECASE,
        ):
            content_start = match.end()
            content_end = html_text.find("</div>", content_start)
            if content_end < 0:
                content_end = min(len(html_text), content_start + 1000)
            blocks.append((match.group(1), html_text[content_start:content_end]))
        return blocks

    @staticmethod
    def _farbnummer_from_color_block(content: str) -> str:
        match = re.search(r"title\s*=\s*['\"]\s*(\d{3})\b", content, re.IGNORECASE)
        return match.group(1) if match else ""

    @staticmethod
    def _color_block_article_url(content: str, base_url: str) -> str:
        match = re.search(
            r"data-article-url\s*=\s*['\"]([^'\"]+)['\"]",
            content,
            re.IGNORECASE,
        )
        if not match:
            return ""
        return urljoin(base_url, unescape(match.group(1)).split("#")[0])

    @staticmethod
    def _purchase_price_source(variant: dict[str, str]) -> str:
        list_price = str(variant.get("data-item-list-price", "")).strip()
        if list_price:
            return list_price
        return str(variant.get("data-item-selling-price", "")).strip()

    @staticmethod
    def _normalize_design(value: str) -> str:
        if value.strip().lower() == "uni":
            return "Uni"
        return value

    @staticmethod
    def _product_name(color_group: str, quality: str) -> str:
        name_parts = ["Venti"]
        if color_group:
            name_parts.append(color_group.lower())
        if quality:
            name_parts.append(quality.lower())
        else:
            name_parts.append("overhemd")
        return " ".join(name_parts)


class CasamodaScraper:
    def __init__(
        self,
        *,
        username: str | None = None,
        password: str | None = None,
        base_dir: str | Path | None = None,
        session: requests.Session | None = None,
        progress_callback: Callable[[str], None] | None = None,
    ):
        self.username = username if username is not None else CASAMODA_USERNAME
        self.password = password if password is not None else CASAMODA_PASSWORD
        self.base_dir = Path(base_dir) if base_dir else Path(BASE_DIR) / "Casamoda"
        self.products_dir = self.base_dir / "products"
        self.logs_dir = self.base_dir / "logs"
        self.price_list_path = self.base_dir / "prijzen.xlsx"
        self.color_map_path = self.base_dir / "kleurcodes.xlsx"
        self.unknown_prices_path = self.base_dir / "unknown_prices.csv"
        self.missing_color_codes_path = self.logs_dir / "missing_color_codes.csv"
        self.autoimport_path = self.base_dir / "autoimport.txt"
        self.progress_callback = progress_callback
        self.session = session or requests.Session()
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36"
                ),
                "Accept-Language": "de-DE,de;q=0.9,nl;q=0.8,en;q=0.7",
            }
        )

    @classmethod
    def scrape_venti(
        cls,
        url: str | None = None,
        *,
        username: str | None = None,
        password: str | None = None,
        progress_callback: Callable[[str], None] | None = None,
    ) -> dict[str, object]:
        scraper = cls(
            username=username,
            password=password,
            progress_callback=progress_callback,
        )
        if url and url.strip():
            return scraper.scrape_category(url.strip())
        return scraper.scrape_autoimport_categories()

    def ensure_autoimport_file(self) -> Path:
        self._ensure_dirs()
        if not self.autoimport_path.exists():
            lines = [
                "# VENTI category URLs. Leave one URL per line.",
                "# The VENTI autoimport uses this file when no category URL is provided.",
                *VENTI_AUTOIMPORT_URLS,
            ]
            self.autoimport_path.write_text(
                "\n".join(lines) + "\n",
                encoding="utf-8",
            )
        return self.autoimport_path

    def resolve_scrape_urls(self, url: str | None) -> list[str]:
        if url and url.strip():
            return [url.strip()]

        self.ensure_autoimport_file()
        seen: set[str] = set()
        urls: list[str] = []
        for line in self.autoimport_path.read_text(encoding="utf-8").splitlines():
            clean_line = line.strip()
            if not clean_line or clean_line.startswith("#"):
                continue
            if clean_line in seen:
                continue
            seen.add(clean_line)
            urls.append(clean_line)
        return urls

    def scrape_autoimport_categories(
        self,
        urls: Iterable[str] | None = None,
    ) -> dict[str, object]:
        self._ensure_dirs()
        scrape_urls = list(urls) if urls is not None else self.resolve_scrape_urls(None)
        if not scrape_urls:
            return {
                "message": "No VENTI autoimport categories configured",
                "error": "No VENTI autoimport categories configured",
                "products": 0,
                "unknown_prices": 0,
                "missing_color_codes": 0,
                "categories": 0,
                "csv_paths": [],
                "all_csv_path": str(
                    self.products_dir
                    / CASAMODA_VENTI_PROFILE.product_aggregate_filename
                ),
            }

        total_products = 0
        total_unknown_prices = 0
        total_missing_color_codes = 0
        csv_paths: list[str] = []

        self._write_unknown_prices([], reset=True)
        self._write_missing_color_codes([], reset=True)
        self._progress(f"Scraping {len(scrape_urls)} VENTI categories...")
        for index, category_url in enumerate(scrape_urls, start=1):
            category_slug = self._category_slug_from_url(category_url)
            self._progress(
                f"Scraping VENTI category {index}/{len(scrape_urls)}: {category_slug}"
            )
            status = self.scrape_category(
                category_url,
                reset_unknown_prices=False,
                reset_missing_color_codes=False,
            )
            total_products += int(status.get("products", 0))
            total_unknown_prices += int(status.get("unknown_prices", 0))
            total_missing_color_codes += int(status.get("missing_color_codes", 0))
            csv_path = str(status.get("csv_path", ""))
            if csv_path:
                csv_paths.append(csv_path)

        all_output_path = self._merge_category_csvs()
        return {
            "message": (
                f"Scraped {total_products} VENTI products from "
                f"{len(scrape_urls)} categories"
            ),
            "error": "" if total_products else "No VENTI products were scraped",
            "products": total_products,
            "unknown_prices": total_unknown_prices,
            "missing_color_codes": total_missing_color_codes,
            "categories": len(scrape_urls),
            "csv_paths": csv_paths,
            "all_csv_path": str(all_output_path),
        }

    def scrape_category(
        self,
        url: str = VENTI_MODERN_FIT_URL,
        *,
        reset_unknown_prices: bool = True,
        reset_missing_color_codes: bool = True,
    ) -> dict[str, str | int]:
        self._ensure_dirs()
        category_slug = self._category_slug_from_url(url)
        self._progress("Loading Casamoda price list...")
        price_list = self._load_price_list()
        color_map = self._load_color_map()
        missing_color_codes: list[CasamodaColorMiss] = []
        missing_color_seen: set[tuple[str, str, str]] = set()

        def add_missing_color(miss: CasamodaColorMiss) -> None:
            key = (miss.article_number, miss.farbnummer, miss.source_url)
            if key in missing_color_seen:
                return
            missing_color_seen.add(key)
            missing_color_codes.append(miss)

        parser = CasamodaParser(
            price_list,
            color_map=color_map,
            missing_color_callback=add_missing_color,
        )
        self._progress("Logging in to Casamoda...")
        self.login()

        self._progress("Collecting VENTI listing pages...")
        listing_urls = self._collect_listing_pages(url)
        detail_urls: list[str] = []
        for page_index, listing_url in enumerate(listing_urls, start=1):
            self._progress(
                f"Reading listing page {page_index}/{len(listing_urls)}..."
            )
            response = self._get(listing_url)
            detail_urls.extend(parser.parse_listing_links(response.text, listing_url))

        detail_queue = list(dict.fromkeys(detail_urls))
        self._progress(f"Found {len(detail_queue)} VENTI color pages to scan.")
        rows: list[dict[str, str]] = []
        seen_skus: set[str] = set()
        scanned_detail_urls: set[str] = set()
        unknown_prices: list[PriceMiss] = []

        while detail_queue:
            detail_url = detail_queue.pop(0)
            if detail_url in scanned_detail_urls:
                continue
            scanned_detail_urls.add(detail_url)
            page_index = len(scanned_detail_urls)
            total_pages = page_index + len(detail_queue)
            self._progress(
                f"Scanning VENTI color page {page_index}/{total_pages}..."
            )
            response = self._get(detail_url)
            for linked_url in parser.parse_article_color_links(response.text, detail_url):
                if linked_url in scanned_detail_urls or linked_url in detail_queue:
                    continue
                detail_queue.append(linked_url)
            selected_farbnummer = parser._selected_farbnummer(response.text, detail_url)
            linked_farbnummers = parser._linked_farbnummer_codes(response.text, detail_url)
            variant_farbnummers = parser._variant_farbnummer_codes(response.text)
            allowed_farbnummers = variant_farbnummers - linked_farbnummers
            if selected_farbnummer:
                allowed_farbnummers.add(selected_farbnummer)
            try:
                product_rows = parser.parse_product_detail(
                    response.text,
                    detail_url,
                    only_farbnummer=allowed_farbnummers or None,
                )
            except UnknownPriceError as ex:
                unknown_prices.extend(self._unknown_prices_from_error(str(ex), detail_url))
                self._progress(
                    f"Skipped color page {page_index}: unknown price found."
                )
                continue
            for row in product_rows:
                if row["sku"] in seen_skus:
                    continue
                seen_skus.add(row["sku"])
                self._apply_category_metadata(row, category_slug, url)
                image_urls = json.loads(row.get("image_urls", "[]"))
                self._progress(f"Downloading images for {row['sku']}...")
                image_count = self._download_images(row["sku"], image_urls)
                row["image_count"] = str(image_count)
                row["has_images"] = str(image_count > 0)
                rows.append(row)
            time.sleep(0.3)

        output_path = self.products_dir / f"{category_slug}.csv"
        if rows:
            pd.DataFrame(rows).to_csv(output_path, index=False, quoting=csv.QUOTE_ALL)
            all_output_path = self._merge_category_csvs()
            self._progress(f"Wrote VENTI CSV: {output_path}")
        else:
            all_output_path = self._merge_category_csvs()

        self._write_unknown_prices(
            unknown_prices,
            category_slug,
            reset=reset_unknown_prices,
        )
        self._write_missing_color_codes(
            missing_color_codes,
            reset=reset_missing_color_codes,
        )
        if unknown_prices:
            self._progress(
                f"Wrote {len(unknown_prices)} unknown price rows to unknown_prices.csv."
            )
        if missing_color_codes:
            self._progress(
                f"Wrote {len(missing_color_codes)} missing color-code rows to logs/missing_color_codes.csv."
            )

        return {
            "message": f"Scraped {len(rows)} VENTI products",
            "error": "" if rows else "No VENTI products were scraped",
            "products": len(rows),
            "unknown_prices": len(unknown_prices),
            "missing_color_codes": len(missing_color_codes),
            "csv_path": str(output_path),
            "all_csv_path": str(all_output_path),
        }

    def login(self) -> None:
        if not self.username or not self.password:
            raise RuntimeError(
                "CASAMODA_USERNAME and CASAMODA_PASSWORD must be set in .env"
            )
        self._get(LOGIN_URL)
        response = self.session.post(
            AUTH_URL,
            data={
                "_username": self.username,
                "_password": self.password,
                "_failure_path": "/de/de/account/login",
            },
            allow_redirects=True,
            timeout=60,
        )
        response.raise_for_status()
        if "account/login" in response.url.lower():
            raise RuntimeError("Casamoda login failed")

    def _collect_listing_pages(self, first_url: str) -> list[str]:
        seen: set[str] = set()
        queued: list[str] = [first_url]

        while queued and len(seen) < 50:
            current_url = queued.pop(0)
            if current_url in seen:
                continue
            seen.add(current_url)
            response = self._get(current_url)
            for href in re.findall(
                r"href\s*=\s*['\"]([^'\"]*page=\d+[^'\"]*)",
                response.text,
            ):
                page_url = urljoin(current_url, href)
                if page_url not in seen and page_url not in queued:
                    queued.append(page_url)

        return sorted(seen, key=self._page_sort_key)

    @staticmethod
    def _page_sort_key(url: str) -> tuple[int, str]:
        match = re.search(r"[?&]page=(\d+)", url)
        return (int(match.group(1)) if match else 1, url)

    def _get(self, url: str, timeout: int = 90) -> requests.Response:
        response = self.session.get(url, timeout=timeout)
        response.raise_for_status()
        return response

    def _load_price_list(self) -> CasamodaPriceList:
        if self.price_list_path.exists():
            return CasamodaPriceList.from_excel(self.price_list_path)

        downloads_copy = Path.home() / "Downloads" / "prijzen.xlsx"
        if downloads_copy.exists():
            return CasamodaPriceList.from_excel(downloads_copy)

        raise FileNotFoundError(
            f"Casamoda price list not found: {self.price_list_path}"
        )

    def _load_color_map(self) -> CasamodaColorMap:
        if self.color_map_path.exists():
            return CasamodaColorMap.from_excel(self.color_map_path)
        return CasamodaColorMap.default()

    def _download_images(self, sku: str, image_urls: list[str]) -> int:
        product_dir = self.products_dir / sku
        product_dir.mkdir(parents=True, exist_ok=True)
        for old_image in product_dir.glob(f"{sku}_*"):
            if old_image.is_file():
                old_image.unlink()
        downloaded = 0
        for index, image_url in enumerate(image_urls):
            extension = self._image_extension(image_url)
            output_path = product_dir / f"{sku}_{index}{extension}"
            try:
                response = self._get(image_url, timeout=20)
            except Exception:
                continue
            if len(response.content) < 1000:
                continue
            output_path.write_bytes(response.content)
            downloaded += 1
        return downloaded

    @staticmethod
    def _image_extension(image_url: str) -> str:
        suffix = Path(urlparse(image_url).path).suffix.lower()
        if suffix in {".jpg", ".jpeg", ".png", ".webp"}:
            return suffix
        return ".jpg"

    def _write_unknown_prices(
        self,
        misses: list[PriceMiss],
        category_slug: str | None = None,
        *,
        reset: bool = False,
    ) -> None:
        fieldnames = [
            "article_number",
            "farbnummer",
            "size",
            "purchase_price",
            "source_url",
        ]
        write_header = reset or not self.unknown_prices_path.exists()
        mode = "w" if reset else "a"
        try:
            with self.unknown_prices_path.open(mode, encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                if write_header:
                    writer.writeheader()
                for miss in misses:
                    writer.writerow(miss.__dict__)
        except PermissionError:
            raise PermissionError(
                f"{self.unknown_prices_path} is locked. Close it and run the scrape again."
            )

    def _write_missing_color_codes(
        self,
        misses: list[CasamodaColorMiss],
        *,
        reset: bool = False,
    ) -> None:
        fieldnames = [
            "article_number",
            "farbnummer",
            "source_url",
        ]
        write_header = reset or not self.missing_color_codes_path.exists()
        mode = "w" if reset else "a"
        try:
            with self.missing_color_codes_path.open(
                mode,
                encoding="utf-8",
                newline="",
            ) as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                if write_header:
                    writer.writeheader()
                for miss in misses:
                    writer.writerow(miss.__dict__)
        except PermissionError:
            raise PermissionError(
                f"{self.missing_color_codes_path} is locked. Close it and run the scrape again."
            )

    @staticmethod
    def _unknown_prices_from_error(error: str, source_url: str) -> list[PriceMiss]:
        misses: list[PriceMiss] = []
        for part in error.split(";"):
            match = re.search(
                r"(?P<article>[A-Za-z0-9-]+)-(?P<color>\d{3})\s+size\s+(?P<size>[^:]+):\s+(?P<price>missing|\d+\.\d{2})",
                part.strip(),
            )
            if not match:
                continue
            misses.append(
                PriceMiss(
                    article_number=match.group("article"),
                    farbnummer=match.group("color"),
                    size=match.group("size").strip(),
                    purchase_price=match.group("price"),
                    source_url=source_url,
                )
            )
        return misses

    def _ensure_dirs(self) -> None:
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self.products_dir.mkdir(parents=True, exist_ok=True)
        self.logs_dir.mkdir(parents=True, exist_ok=True)

    def _progress(self, message: str) -> None:
        if self.progress_callback is not None:
            self.progress_callback(message)

    @staticmethod
    def _category_slug_from_url(url: str) -> str:
        path_name = Path(urlparse(url).path).name or "venti"
        path_name = path_name.replace("product-list-", "")
        slug = re.sub(r"[^A-Za-z0-9]+", "_", path_name).strip("_").lower()
        return slug or "venti"

    @staticmethod
    def _apply_category_metadata(
        row: dict[str, str],
        category_slug: str,
        category_url: str,
    ) -> None:
        row["source_category_slug"] = category_slug
        row["source_category_url"] = category_url
        row["magento_ready"] = "False"
        row["blocked_reason"] = ""

        if category_slug in {
            "venti_modern_fit",
            "venti_body_fit",
            "venti_comfort_fit",
            "venti_jerseyflex",
            "venti_evening",
        }:
            row["category"] = "shirts"
            row["magento_ready"] = "True"
            CasamodaScraper._block_missing_color(row)
            return

        if category_slug == "venti_polos_shirts":
            product_name = str(row.get("name", "")).lower()
            if "polo" in product_name:
                row["category"] = "polos"
                row["blocked_reason"] = (
                    "Polos/Shirts category is scraped for review only; "
                    "Magento polo mapping is not approved yet."
                )
            elif "t-shirt" in product_name or "tanktop" in product_name:
                row["category"] = "t-shirts"
                row["fit"] = "Modern Fit"
                row["magento_ready"] = "True"
                CasamodaScraper._block_missing_color(row)
                return
            else:
                row["category"] = "review"
                row["blocked_reason"] = (
                    "Polos/Shirts category is scraped for review only; "
                    "Magento mapping is not approved yet."
                )
            CasamodaScraper._block_missing_color(row)
            return

        row["category"] = "review"
        row["blocked_reason"] = (
            f"Category {category_slug} is scraped for review only; "
            "Magento mapping is not approved yet."
        )
        CasamodaScraper._block_missing_color(row)

    @staticmethod
    def _block_missing_color(row: dict[str, str]) -> None:
        if str(row.get("color_missing", "")).strip().lower() != "true":
            return

        row["magento_ready"] = "False"
        farbnummer = str(row.get("farbnummer", "")).strip() or "unknown"
        reason = f"Color code {farbnummer} is missing from kleurcodes.xlsx."
        existing_reason = row.get("blocked_reason", "")
        row["blocked_reason"] = (
            f"{existing_reason} {reason}".strip() if existing_reason else reason
        )

    def _merge_category_csvs(self) -> Path:
        output_path = (
            self.products_dir / CASAMODA_VENTI_PROFILE.product_aggregate_filename
        )
        csv_files = sorted(
            path
            for path in self.products_dir.glob("venti_*.csv")
            if path.name not in {"all.csv", "venti_all.csv"}
        )
        frames: list[pd.DataFrame] = []
        for csv_file in csv_files:
            try:
                frame = pd.read_csv(csv_file)
            except Exception:
                continue
            if not frame.empty:
                frames.append(frame)

        if not frames:
            empty_frame = pd.DataFrame(columns=CASAMODA_PRODUCT_FIELDS)
            try:
                empty_frame.to_csv(output_path, index=False, quoting=csv.QUOTE_ALL)
                self._remove_legacy_all_csv()
                return output_path
            except PermissionError:
                fallback_path = (
                    self.products_dir / f"venti_all_{int(time.time())}.csv"
                )
                empty_frame.to_csv(fallback_path, index=False, quoting=csv.QUOTE_ALL)
                self._remove_legacy_all_csv()
                self._progress(
                    f"{output_path.name} is locked; wrote empty merged VENTI CSV to {fallback_path.name} instead."
                )
                return fallback_path

        merged = pd.concat(frames, ignore_index=True)
        if "sku" in merged.columns:
            merged = merged.drop_duplicates(subset=["sku"], keep="last")
        try:
            merged.to_csv(output_path, index=False, quoting=csv.QUOTE_ALL)
            self._remove_legacy_all_csv()
            return output_path
        except PermissionError:
            fallback_path = (
                self.products_dir / f"venti_all_{int(time.time())}.csv"
            )
            merged.to_csv(fallback_path, index=False, quoting=csv.QUOTE_ALL)
            self._remove_legacy_all_csv()
            self._progress(
                f"{output_path.name} is locked; wrote merged VENTI CSV to {fallback_path.name} instead."
            )
            return fallback_path

    def _remove_legacy_all_csv(self) -> None:
        legacy_path = self.products_dir / "all.csv"
        if not legacy_path.exists():
            return
        try:
            legacy_path.unlink()
        except PermissionError:
            self._progress(
                "Could not remove legacy Casamoda all.csv because it is locked."
            )
